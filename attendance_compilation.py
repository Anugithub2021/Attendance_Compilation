import os # import os module
os.system("cls") 
os.chdir(r'C:\Users\User\Documents\GitHub\Attendance_Compilation') # active directory to current directory by providing the location

# importing modules for mailing
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# importing modules for file handling
import pandas as pd #panda module
from datetime import datetime
import openpyxl as op

def send_mail(fromaddr, frompasswd, toaddr, msg_subject, msg_body, file_path): 
    try: # creating message object
        msg = MIMEMultipart()
        print("[+] Message Object Created")
    except:
        print("[-] Error in Creating Message Object")
        return

    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = msg_subject

    body = msg_body
    msg.attach(MIMEText(body, 'plain'))

    filename = file_path
    attachment = open(filename, "rb")

    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('body-Disposition', "attachment; filename= %s" % filename)

    try: # attaching file
        msg.attach(p)
        print("[+] File Attached")
    except:
        print("[-] Error in Attaching file")
        return
    
    try: # creating SMTP session
        #s = smtplib.SMTP('smtp.gmail.com', 587)
        s = smtplib.SMTP('stud.iitp.ac.in', 587)
        print("[+] SMTP Session Created")
    except:
        print("[-] Error in creating SMTP session")
        return
    try:
        s.starttls() # starting connection
        print("Started the connection")
    except Exception as e:
        print(e)
        print("Weren't able to start the connection")
        return

    try: # login in the account
        s.login(fromaddr, frompasswd)
        print("[+] Login Successful")
    except:
        print("[-] Login Failed")

    text = msg.as_string()

    try: # sending mail
        s.sendmail(fromaddr, toaddr, text)
        print("[+] Mail Sent successfully")
    except:
        print('[-] Mail not sent')

    s.quit() # ending the session

def attendance_report(): # last function call
    # reading input files
    reader=pd.read_csv(r'input_registered_students.csv') 
    inp_att=pd.read_csv(r'input_attendance.csv') 

    try:
        reader.dropna() # doing this so that we don't get any key error and then finding the number of registered students
        inp_att=inp_att.dropna() # dropping null rows
    except:
        print("Can't drop the null values")
    
    lenth=len(reader)

    map={} # mapping roll number with their names
    try:
        for i in range(lenth): map[reader.iloc[i][0]]=reader.iloc[i][1]
    except:
        print("Unable to map the roll number and names")
    
    inp_att["Timestamp"]=pd.to_datetime(inp_att["Timestamp"],dayfirst=1) # converting string to datetime format
    
    list_roll=list(reader["Roll No"])
    date_time=list(inp_att["Timestamp"])

    try:
        period=["14:00","15:00"]
        period[0]=datetime.strptime(period[0],"%H:%M")
        period[1]=datetime.strptime(period[1],"%H:%M")
    except:
        print("Unable to set the period.")

    mark_roll=[] # a list of only the marked roll numbers
    for i in inp_att["Attendance"]:
        temp=i.split()
        mark_roll.append(temp[0])

    lec_taken=[]
    tot_cnt={}
    inv_cnt={}

    x=0
    for i in date_time:
        if i.isoweekday()==1 or i.isoweekday()==4: # isoweekday gives 1 for monday and 4 for thursday
            if x!=i.date(): # no repetition
                tot_cnt[i.date()]={} # initialization
                inv_cnt[i.date()]={} # initialization
                for j in list_roll:
                    tot_cnt[i.date()][j]=0 # initialization
                    inv_cnt[i.date()][j]=0 # initialization

                lec_taken.append(i.date())

            x=i.date()

    no_of_lec=len(lec_taken)

    for i in list_roll: # for every roll number individually
        tot=0
        inv=0
        x=0

        for j in range(len(mark_roll)): # for complete attendance sheet
            if i==mark_roll[j]:
                if date_time[j].isoweekday()==1 or date_time[j].isoweekday()==4:
                    if x!=date_time[j].date(): # for different dates
                        if x!=0:
                            tot_cnt[x][i]=tot
                            inv_cnt[x][i]=inv
                        tot=0
                        inv=0
                        x=date_time[j].date()

                    tot+=1
                    try:
                        if date_time[j].time()<period[0].time() or date_time[j].time()>period[1].time(): inv+=1
                    except:
                        print("Failed in comparing time period")
        
        if x!=0:
            tot_cnt[x][i]=tot
            inv_cnt[x][i]=inv
        
    os.chdir(r'C:\Users\User\Documents\GitHub\Attendance_Compilation\output')

    last={}
    for i in list_roll: # for every student
        last[i]=[]
        wb=op.Workbook()
        sheet=wb.active
        sheet.cell(row=1,column=1).value="Date"
        sheet.cell(row=1,column=2).value="Roll"
        sheet.cell(row=1,column=3).value="Name"
        sheet.cell(row=1,column=4).value="Attendance"
        sheet.cell(row=1,column=5).value="Real"
        sheet.cell(row=1,column=6).value="Duplicate"
        sheet.cell(row=1,column=7).value="Invalid"
        sheet.cell(row=1,column=8).value="Absent"
        sheet.cell(row=2,column=2).value=i
        sheet.cell(row=2,column=3).value=map[i]

        for j in range(len(lec_taken)):
            tot=tot_cnt[lec_taken[j]][i] 
            inv=inv_cnt[lec_taken[j]][i]

            sheet.cell(row=j+3,column=1).value=lec_taken[j]
            sheet.cell(row=j+3,column=4).value=tot

            if tot-inv>0:
                sheet.cell(row=j+3,column=5).value=1
                sheet.cell(row=j+3,column=6).value=tot-inv-1 
                sheet.cell(row=j+3,column=8).value=0
                last[i].append('P')
            else:
                sheet.cell(row=j+3,column=5).value=0 
                sheet.cell(row=j+3,column=6).value=0
                sheet.cell(row=j+3,column=8).value=1
                last[i].append('A')

            sheet.cell(row=j+3,column=7).value=inv
        wb.save(filename=i+".xlsx")

    wb=op.Workbook()
    sheet=wb.active
    sheet.cell(row=1,column=1).value="Roll"
    sheet.cell(row=1,column=2).value="Name"

    for i in range(len(lec_taken)):
        sheet.cell(row=1,column=i+3).value=lec_taken[i] 
    
    sheet.cell(row=1,column=3+len(lec_taken)).value="Actual Lecture Taken"
    sheet.cell(row=1,column=4+len(lec_taken)).value="Total Real"
    sheet.cell(row=1,column=5+len(lec_taken)).value="% Attendance"

    for i in range(lenth):
        sheet.cell(row=i+2,column=1).value=i
        sheet.cell(row=i+2,column=2).value=map[list_roll[i]]

        col=3; cnt=0 
        for j in last[list_roll[i]]:
            sheet.cell(row=i+2,column=col).value=j
            if(j=='P'):
                cnt+=1
            col+=1
        sheet.cell(row=i+2,column=col).value=no_of_lec
        sheet.cell(row=i+2,column=col+1).value=cnt
        sheet.cell(row=i+2,column=col+2).value=round((cnt/no_of_lec)*100,2) 
    
    wb.save(filename="attendance_report_consolidated.xlsx")

    FROM_ADDR = "anuradha_2001cb10@iitp.ac.in"
    FROM_PASSWD = "changeme"
    receiver="changeme"

    Subject="Attendance Report"
    body="Anuradha Das Group 2001CB10" 
    file_path="attendance_report_consolidated.xlsx"
    try:
        send_mail(FROM_ADDR, FROM_PASSWD, receiver, Subject, body, file_path)
    except:
        print("Weren't able to send the mail")
    
attendance_report() # call the function